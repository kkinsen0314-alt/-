#!/usr/bin/env python3
"""Process exported Feice data for the Axing project."""

from __future__ import annotations

import argparse
import re
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl import Workbook


EXCLUDED_ORDER_STATUSES = {"订单已取消", "售后取消"}
try:
    SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
except Exception:
    SHANGHAI_TZ = None


def normalize_phone(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    elif isinstance(value, int):
        text = str(value)
    else:
        text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if len(digits) == 13 and digits.startswith("86"):
        digits = digits[-11:]
    return digits or text


def shanghai_today() -> date:
    if SHANGHAI_TZ is not None:
        return datetime.now(SHANGHAI_TZ).date()
    return (datetime.utcnow() + timedelta(hours=8)).date()


def is_blank_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return not any(value not in (None, "") for value in values)


def resolve_input_path(
    explicit_path: str | Path | None,
    download_dir: str | Path | None,
    patterns: tuple[str, ...],
    label: str,
) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"找不到{label}：{path}")
        return path
    if not download_dir:
        raise ValueError(f"请提供{label}路径，或使用 --download-dir 自动选择最新导出文件")

    root = Path(download_dir).expanduser().resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"下载目录不存在：{root}")
    candidates = {
        path
        for pattern in patterns
        for path in root.glob(pattern)
        if path.is_file()
        and not path.name.endswith(".crdownload")
        and "阿星进量" not in path.stem
    }
    if not candidates:
        joined = "、".join(patterns)
        raise FileNotFoundError(f"{root} 中没有找到{label}（匹配：{joined}）")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def header_index(headers: list[Any], name: str, path: Path) -> int:
    for index, value in enumerate(headers):
        if str(value or "").strip() == name:
            return index
    raise ValueError(f"文件缺少必要列“{name}”：{path}")


def parse_time(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return datetime.max


def copy_cell(source: Any, target: Any) -> None:
    target.value = source.value
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)


def copy_sheet_settings(source: Any, target: Any) -> None:
    target.freeze_panes = source.freeze_panes
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden


def process_axing_jinliang(
    clue_path: str | Path,
    order_path: str | Path,
    output_dir: str | Path | None = None,
    output_date: date | None = None,
) -> dict[str, Any]:
    clue_path = Path(clue_path).expanduser().resolve()
    order_path = Path(order_path).expanduser().resolve()
    if not clue_path.is_file():
        raise FileNotFoundError(f"找不到线索表：{clue_path}")
    if not order_path.is_file():
        raise FileNotFoundError(f"找不到订单表：{order_path}")

    target_dir = Path(output_dir).expanduser().resolve() if output_dir else clue_path.parent
    target_dir.mkdir(parents=True, exist_ok=True)
    output_date = output_date or shanghai_today()

    order_book = openpyxl.load_workbook(order_path, data_only=True)
    try:
        order_sheet = order_book.active
        order_headers = [order_sheet.cell(1, col).value for col in range(1, order_sheet.max_column + 1)]
        contact_col = header_index(order_headers, "联系方式", order_path)
        status_col = header_index(order_headers, "订单状态", order_path)
        product_name_col = header_index(order_headers, "商品名称", order_path)
        order_map: dict[str, Any] = {}
        excluded_count = 0
        duplicate_phone_count = 0
        blank_order_count = 0
        for row in order_sheet.iter_rows(min_row=2, values_only=True):
            if is_blank_row(row):
                blank_order_count += 1
                continue
            status = str(row[status_col] or "").strip()
            if status in EXCLUDED_ORDER_STATUSES:
                excluded_count += 1
                continue
            phone = normalize_phone(row[contact_col])
            if phone:
                if phone in order_map:
                    duplicate_phone_count += 1
                else:
                    order_map[phone] = row[product_name_col]
    finally:
        order_book.close()

    clue_book = openpyxl.load_workbook(clue_path, data_only=False)
    try:
        clue_sheet = clue_book.active
        clue_headers = [clue_sheet.cell(1, col).value for col in range(1, clue_sheet.max_column + 1)]
        phone_col = header_index(clue_headers, "手机号码", clue_path)
        form_time_col = header_index(clue_headers, "表单填写时间", clue_path)
        wechat_col = header_index(clue_headers, "微信号", clue_path)
        qq_col = next((index for index, value in enumerate(clue_headers) if str(value or "").strip() == "QQ号"), None)
        source_columns = [index for index in range(clue_sheet.max_column) if index != qq_col]
        records: list[tuple[tuple[Any, ...], int]] = []
        matched_count = 0
        unmatched_count = 0
        blank_clue_count = 0
        for row_number, row in enumerate(clue_sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row]
            if is_blank_row(values):
                blank_clue_count += 1
                continue
            phone = normalize_phone(values[phone_col])
            if phone not in order_map:
                unmatched_count += 1
                continue
            values[wechat_col] = order_map[phone]
            records.append((tuple(values), row_number))
            matched_count += 1
        records.sort(key=lambda item: parse_time(item[0][form_time_col]))

        result_book = Workbook()
        result_sheet = result_book.active
        result_sheet.title = clue_sheet.title or "线索"
        copy_sheet_settings(clue_sheet, result_sheet)
        for output_col, source_col in enumerate(source_columns, start=1):
            copy_cell(clue_sheet.cell(1, source_col + 1), result_sheet.cell(1, output_col))
        for output_row, (values, source_row) in enumerate(records, start=2):
            for output_col, source_col in enumerate(source_columns, start=1):
                target = result_sheet.cell(output_row, output_col)
                target.value = values[source_col]
                source_cell = clue_sheet.cell(source_row, source_col + 1)
                if source_cell.has_style:
                    target._style = copy(source_cell._style)
                if source_cell.number_format:
                    target.number_format = source_cell.number_format
        if result_sheet.max_row >= 1 and result_sheet.max_column >= 1:
            result_sheet.auto_filter.ref = result_sheet.dimensions
        output_path = target_dir / f"线索管理明细表_阿星进量_{output_date:%Y-%m-%d}.xlsx"
        temp_output_path = output_path.with_name(f".{output_path.stem}.tmp.xlsx")
        try:
            result_book.save(temp_output_path)
            result_book.close()
            temp_output_path.replace(output_path)
        except Exception:
            result_book.close()
            if temp_output_path.exists():
                temp_output_path.unlink()
            raise
    finally:
        clue_book.close()

    return {
        "total_clues": matched_count + unmatched_count,
        "matched_count": matched_count,
        "unmatched_count": unmatched_count,
        "excluded_order_count": excluded_count,
        "duplicate_phone_count": duplicate_phone_count,
        "blank_order_count": blank_order_count,
        "blank_clue_count": blank_clue_count,
        "qq_column_removed": qq_col is not None,
        "output_path": str(output_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="处理阿星项目线索和店铺订单导出文件")
    parser.add_argument("--clue", help="线索管理明细表路径；省略时从 --download-dir 自动选择最新文件")
    parser.add_argument("--orders", help="店铺订单信息路径；省略时从 --download-dir 自动选择最新文件")
    parser.add_argument("--download-dir", help="飞策导出文件目录，自动选择最新线索表和订单表")
    parser.add_argument("--output-dir", help="输出目录，默认使用线索表所在目录")
    parser.add_argument("--date", dest="output_date", help="输出日期，格式 YYYY-MM-DD，默认使用今天")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.download_dir and (not args.clue or not args.orders):
        raise SystemExit("请同时提供 --clue 和 --orders，或提供 --download-dir")
    clue_path = resolve_input_path(
        args.clue,
        args.download_dir,
        ("线索管理明细表*.xlsx",),
        "线索表",
    )
    order_path = resolve_input_path(
        args.orders,
        args.download_dir,
        ("店铺订单*.xlsx", "店铺订单信息*.xlsx"),
        "订单表",
    )
    output_date = date.fromisoformat(args.output_date) if args.output_date else shanghai_today()
    result = process_axing_jinliang(clue_path, order_path, args.output_dir, output_date)
    print(f"线索文件: {clue_path}")
    print(f"订单文件: {order_path}")
    print(f"线索总数: {result['total_clues']}")
    print(f"匹配成功: {result['matched_count']}")
    print(f"未匹配并删除: {result['unmatched_count']}")
    print(f"排除取消订单: {result['excluded_order_count']}")
    print(f"重复订单手机号: {result['duplicate_phone_count']}")
    print(f"跳过空订单行: {result['blank_order_count']}")
    print(f"跳过空线索行: {result['blank_clue_count']}")
    print(f"QQ号列已删除: {'是' if result['qq_column_removed'] else '否'}")
    print(f"输出文件: {result['output_path']}")


if __name__ == "__main__":
    main()
